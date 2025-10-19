"""
Report Generation Module
Generate professional reports in multiple formats:
- PDF reports with charts
- Excel spreadsheets
- JSON exports
- HTML reports
"""

from typing import Dict, Any, List, Optional
from datetime import datetime
import json
import io
import base64


class ReportGenerator:
    """Generate reports in various formats"""
    
    def __init__(self):
        self.templates = {}
    
    def generate_pdf_report(self, data: Dict[str, Any], report_type: str = "investigation") -> Dict[str, Any]:
        """
        Generate PDF report
        Note: In production, use libraries like ReportLab or WeasyPrint
        This is a placeholder that returns report metadata
        """
        report = {
            "format": "PDF",
            "report_type": report_type,
            "generated_at": datetime.now().isoformat(),
            "status": "generated",
            "note": "PDF generation requires reportlab library (pip install reportlab)"
        }
        
        try:
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.lib import colors
            from reportlab.lib.enums import TA_CENTER, TA_LEFT
            
            # Create PDF in memory
            buffer = io.BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=letter)
            styles = getSampleStyleSheet()
            story = []
            
            # Title
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                textColor=colors.HexColor('#a855f7'),
                spaceAfter=30,
                alignment=TA_CENTER
            )
            
            story.append(Paragraph("🛡️ EthGuardian AI", title_style))
            story.append(Paragraph(f"{report_type.upper()} REPORT", title_style))
            story.append(Spacer(1, 0.5*inch))
            
            # Report metadata
            metadata_data = [
                ['Report ID:', data.get('report_id', 'N/A')],
                ['Generated:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
                ['Address:', data.get('address', 'N/A')],
                ['Risk Score:', f"{data.get('risk_score', 0)}/100"]
            ]
            
            metadata_table = Table(metadata_data, colWidths=[2*inch, 4*inch])
            metadata_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f0f0ff')),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey)
            ]))
            
            story.append(metadata_table)
            story.append(Spacer(1, 0.5*inch))
            
            # Executive Summary
            story.append(Paragraph("Executive Summary", styles['Heading2']))
            story.append(Spacer(1, 0.2*inch))
            summary_text = data.get('summary', 'No summary available.')
            story.append(Paragraph(summary_text, styles['Normal']))
            story.append(Spacer(1, 0.3*inch))
            
            # Risk Assessment
            story.append(Paragraph("Risk Assessment", styles['Heading2']))
            story.append(Spacer(1, 0.2*inch))
            
            risk_data = [
                ['Metric', 'Score', 'Status'],
                ['Overall Risk', str(data.get('risk_score', 0)), self._get_risk_status(data.get('risk_score', 0))],
                ['Pattern Score', str(data.get('pattern_score', 0)), self._get_risk_status(data.get('pattern_score', 0))],
                ['Fraud Score', str(data.get('fraud_score', 0)), self._get_risk_status(data.get('fraud_score', 0))],
                ['Analytics Score', str(data.get('analytics_score', 0)), self._get_risk_status(data.get('analytics_score', 0))]
            ]
            
            risk_table = Table(risk_data, colWidths=[2*inch, 1.5*inch, 1.5*inch])
            risk_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#a855f7')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(risk_table)
            story.append(Spacer(1, 0.3*inch))
            
            # Findings
            if 'findings' in data:
                story.append(PageBreak())
                story.append(Paragraph("Detailed Findings", styles['Heading2']))
                story.append(Spacer(1, 0.2*inch))
                
                findings_text = json.dumps(data['findings'], indent=2)
                story.append(Paragraph(f"<pre>{findings_text[:2000]}</pre>", styles['Code']))
            
            # Build PDF
            doc.build(story)
            pdf_bytes = buffer.getvalue()
            buffer.close()
            
            report["pdf_data"] = base64.b64encode(pdf_bytes).decode('utf-8')
            report["size_bytes"] = len(pdf_bytes)
            report["status"] = "success"
            
        except ImportError:
            report["status"] = "library_not_available"
            report["message"] = "Install reportlab: pip install reportlab"
        except Exception as e:
            report["status"] = "error"
            report["error"] = str(e)
        
        return report
    
    def generate_excel_report(self, data: Dict[str, Any], report_type: str = "investigation") -> Dict[str, Any]:
        """
        Generate Excel report
        Note: In production, use libraries like openpyxl or xlsxwriter
        """
        report = {
            "format": "EXCEL",
            "report_type": report_type,
            "generated_at": datetime.now().isoformat(),
            "status": "generated",
            "note": "Excel generation requires openpyxl library (pip install openpyxl)"
        }
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            
            wb = Workbook()
            
            # Summary Sheet
            ws_summary = wb.active
            ws_summary.title = "Summary"
            
            # Header
            ws_summary['A1'] = "EthGuardian AI - Investigation Report"
            ws_summary['A1'].font = Font(size=16, bold=True, color="a855f7")
            ws_summary['A1'].alignment = Alignment(horizontal='center')
            ws_summary.merge_cells('A1:D1')
            
            # Metadata
            row = 3
            metadata_fields = [
                ('Report ID', data.get('report_id', 'N/A')),
                ('Generated', datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
                ('Address', data.get('address', 'N/A')),
                ('Risk Score', f"{data.get('risk_score', 0)}/100"),
            ]
            
            for label, value in metadata_fields:
                ws_summary[f'A{row}'] = label
                ws_summary[f'A{row}'].font = Font(bold=True)
                ws_summary[f'B{row}'] = value
                row += 1
            
            # Risk Scores Sheet
            ws_risk = wb.create_sheet("Risk Scores")
            ws_risk['A1'] = "Risk Assessment"
            ws_risk['A1'].font = Font(size=14, bold=True)
            
            headers = ['Metric', 'Score', 'Status']
            for col, header in enumerate(headers, 1):
                cell = ws_risk.cell(row=3, column=col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="a855f7", end_end_color="a855f7", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
            
            risk_scores = [
                ('Overall Risk', data.get('risk_score', 0)),
                ('Pattern Score', data.get('pattern_score', 0)),
                ('Fraud Score', data.get('fraud_score', 0)),
                ('Analytics Score', data.get('analytics_score', 0))
            ]
            
            for idx, (metric, score) in enumerate(risk_scores, 4):
                ws_risk[f'A{idx}'] = metric
                ws_risk[f'B{idx}'] = score
                ws_risk[f'C{idx}'] = self._get_risk_status(score)
            
            # Transactions Sheet (if available)
            if 'transactions' in data:
                ws_tx = wb.create_sheet("Transactions")
                ws_tx['A1'] = "Transaction History"
                ws_tx['A1'].font = Font(size=14, bold=True)
                
                tx_headers = ['Timestamp', 'From', 'To', 'Value (ETH)', 'Type']
                for col, header in enumerate(tx_headers, 1):
                    cell = ws_tx.cell(row=3, column=col)
                    cell.value = header
                    cell.font = Font(bold=True)
                
                for idx, tx in enumerate(data['transactions'][:100], 4):  # Limit to 100
                    ws_tx[f'A{idx}'] = tx.get('timestamp', 'N/A')
                    ws_tx[f'B{idx}'] = tx.get('from', 'N/A')
                    ws_tx[f'C{idx}'] = tx.get('to', 'N/A')
                    ws_tx[f'D{idx}'] = tx.get('value', 0)
                    ws_tx[f'E{idx}'] = tx.get('type', 'N/A')
            
            # Save to bytes
            buffer = io.BytesIO()
            wb.save(buffer)
            excel_bytes = buffer.getvalue()
            buffer.close()
            
            report["excel_data"] = base64.b64encode(excel_bytes).decode('utf-8')
            report["size_bytes"] = len(excel_bytes)
            report["status"] = "success"
            
        except ImportError:
            report["status"] = "library_not_available"
            report["message"] = "Install openpyxl: pip install openpyxl"
        except Exception as e:
            report["status"] = "error"
            report["error"] = str(e)
        
        return report
    
    def generate_json_report(self, data: Dict[str, Any]) -> Dict[str, Any]:
        """Generate JSON report"""
        report = {
            "format": "JSON",
            "generated_at": datetime.now().isoformat(),
            "data": data
        }
        
        # Pretty print JSON
        report["json_string"] = json.dumps(data, indent=2, default=str)
        report["size_bytes"] = len(report["json_string"])
        report["status"] = "success"
        
        return report
    
    def generate_html_report(self, data: Dict[str, Any]) -> Dict[str, Any]:
        """Generate HTML report"""
        html = f"""
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>EthGuardian AI Report</title>
    <style>
        * {{ box-sizing: border-box; }}
        body {{
            font-family: 'Inter', -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif;
            background: linear-gradient(135deg, #0a0118 0%, #1a0b2e 50%, #0f051d 100%);
            color: #f0f0ff;
            margin: 0;
            padding: 20px;
        }}
        .container {{
            max-width: 1200px;
            margin: 0 auto;
            background: rgba(20, 10, 40, 0.8);
            border-radius: 16px;
            padding: 40px;
            box-shadow: 0 8px 32px rgba(0, 0, 0, 0.4);
        }}
        .header {{
            text-align: center;
            margin-bottom: 40px;
            padding-bottom: 20px;
            border-bottom: 2px solid rgba(168, 85, 247, 0.3);
        }}
        h1 {{
            font-size: 36px;
            background: linear-gradient(90deg, #a855f7, #06b6d4);
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
            margin: 0;
        }}
        .badge {{
            display: inline-block;
            padding: 8px 16px;
            border-radius: 999px;
            background: linear-gradient(135deg, #a855f7, #06b6d4);
            color: white;
            font-weight: 700;
            margin: 10px 0;
        }}
        .section {{
            margin: 30px 0;
            padding: 20px;
            background: rgba(13, 10, 25, 0.6);
            border-radius: 12px;
            border: 1px solid rgba(168, 85, 247, 0.2);
        }}
        .section h2 {{
            color: #a855f7;
            margin-top: 0;
        }}
        .metric {{
            display: grid;
            grid-template-columns: 200px 1fr;
            gap: 20px;
            padding: 12px 0;
            border-bottom: 1px solid rgba(168, 85, 247, 0.1);
        }}
        .metric:last-child {{
            border-bottom: none;
        }}
        .metric-label {{
            color: #a78bfa;
            font-weight: 600;
        }}
        .metric-value {{
            color: #f0f0ff;
        }}
        .risk-high {{ color: #f43f5e; }}
        .risk-medium {{ color: #FFA94D; }}
        .risk-low {{ color: #10b981; }}
        table {{
            width: 100%;
            border-collapse: collapse;
            margin: 20px 0;
        }}
        th {{
            background: linear-gradient(135deg, #a855f7, #06b6d4);
            color: white;
            padding: 12px;
            text-align: left;
        }}
        td {{
            padding: 10px 12px;
            border-bottom: 1px solid rgba(168, 85, 247, 0.1);
        }}
        code {{
            background: rgba(0, 0, 0, 0.4);
            padding: 2px 6px;
            border-radius: 4px;
            font-family: 'Monaco', monospace;
        }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>🛡️ EthGuardian AI</h1>
            <div class="badge">Investigation Report</div>
            <p>Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
        </div>
        
        <div class="section">
            <h2>Report Summary</h2>
            <div class="metric">
                <div class="metric-label">Report ID:</div>
                <div class="metric-value">{data.get('report_id', 'N/A')}</div>
            </div>
            <div class="metric">
                <div class="metric-label">Address:</div>
                <div class="metric-value"><code>{data.get('address', 'N/A')}</code></div>
            </div>
            <div class="metric">
                <div class="metric-label">Risk Score:</div>
                <div class="metric-value {self._get_risk_class(data.get('risk_score', 0))}">{data.get('risk_score', 0)}/100</div>
            </div>
        </div>
        
        <div class="section">
            <h2>Risk Assessment</h2>
            <table>
                <tr>
                    <th>Metric</th>
                    <th>Score</th>
                    <th>Status</th>
                </tr>
                <tr>
                    <td>Overall Risk</td>
                    <td>{data.get('risk_score', 0)}</td>
                    <td class="{self._get_risk_class(data.get('risk_score', 0))}">{self._get_risk_status(data.get('risk_score', 0))}</td>
                </tr>
                <tr>
                    <td>Pattern Score</td>
                    <td>{data.get('pattern_score', 0)}</td>
                    <td class="{self._get_risk_class(data.get('pattern_score', 0))}">{self._get_risk_status(data.get('pattern_score', 0))}</td>
                </tr>
                <tr>
                    <td>Fraud Score</td>
                    <td>{data.get('fraud_score', 0)}</td>
                    <td class="{self._get_risk_class(data.get('fraud_score', 0))}">{self._get_risk_status(data.get('fraud_score', 0))}</td>
                </tr>
                <tr>
                    <td>Analytics Score</td>
                    <td>{data.get('analytics_score', 0)}</td>
                    <td class="{self._get_risk_class(data.get('analytics_score', 0))}">{self._get_risk_status(data.get('analytics_score', 0))}</td>
                </tr>
            </table>
        </div>
        
        <div class="section">
            <h2>Findings</h2>
            <pre style="background: rgba(0,0,0,0.3); padding: 20px; border-radius: 8px; overflow-x: auto;">
{json.dumps(data.get('findings', {}), indent=2)}
            </pre>
        </div>
        
        <div class="section">
            <p style="text-align: center; color: #a78bfa; font-size: 12px;">
                Generated by EthGuardian AI - Ethereum AML Monitoring Platform
            </p>
        </div>
    </div>
</body>
</html>
"""
        
        report = {
            "format": "HTML",
            "generated_at": datetime.now().isoformat(),
            "html": html,
            "size_bytes": len(html),
            "status": "success"
        }
        
        return report
    
    def _get_risk_status(self, score: int) -> str:
        """Get risk status text"""
        if score >= 70:
            return "HIGH RISK"
        elif score >= 40:
            return "MEDIUM RISK"
        else:
            return "LOW RISK"
    
    def _get_risk_class(self, score: int) -> str:
        """Get CSS class for risk level"""
        if score >= 70:
            return "risk-high"
        elif score >= 40:
            return "risk-medium"
        else:
            return "risk-low"


# Global report generator
report_generator = ReportGenerator()

